from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.models.employee import Employee


# Maps Employee field name -> raw Excel column name (as pandas parses it,
# with .1 suffixes for duplicate headers in the source sheet)
COLUMN_MAP = {
    "epf": "EPF",
    "name": "Name",
    "department": "Name.1",  # 2nd "Name" column in the sheet is actually Department

    "basic_salary": "BRA+ Anodizing+Salary",
    "holiday_pay": "B Holiday",
    "no_pay": "No pay deduction",
    "nopay_revise": "Nopay Revise",
    "salary_adjustment": "OT Arreares(Salary Adj)",

    "night_shift": "Night Shift",
    "saturday_night": "Saturday",
    "sunday_payment": "Sunday (Off)",
    "poya_mercantile": "Poya+ Mer",
    "mercantile_special_payment": "Mercantile Special Payment",
    "refund_iceu_amount": "Refund ICEU Amount",
    "ot_arrears": "OT arreys",
    "salary_arrears_deduction": "Salary arrears",
    "over_time": "OT",
    "double_time": "DOT",
    "powder_incentive": "powder incentive",
    "tailoring_fee": "Tailoring Fee",
    "att_incentive": "Attendance %",
    "melt_incentive": "Melt Allowance",
    "team_leader_incentive": "Team Leader Allowance",
    "chemical_incentive": "Chemical Allowance",
    "export_incentive": "Export Loading Allowances",
    "production_incentive": "Production %",

    "salary_advance": "Salary Advance",
    "festival_advance": "festival advance",
    "vision_care": "Vison Care Amount",
    "meals": "Meals",
    "welfare_society": "welfare",
    "other_deduction_tp": "Other deduction (TP & Mobile phone & Bonus Adv )",
    "festival_advance_guarantors": "Festival Advance - claim from guarantors",
    "iceu_member_fee": "ICEU Member fee",
    "paye_tax": "Paye tax",
    "stamp_duty": "Stamp duty",

    "casual_leave": "Casual Leaves days",
    "rate_per_not": "Rate N.o.t",
    "rate_per_dot": "Rate per D.O.t",
}


class ExcelReader:
    def __init__(self, input_folder=None):
        from src.config import INPUT_DIR
        self.input_folder = Path(input_folder) if input_folder else INPUT_DIR

    def load_excel(self, filename):
        file_path = self.input_folder / filename

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if file_path.suffix.lower() == ".csv":
            return pd.read_csv(file_path, header=1)

        return pd.read_excel(file_path, header=1)

    def get_pay_period(self, filename):
        """Reads the title row (e.g. 'Payslip - Month of June 2026') and
        extracts 'June 2026'. Falls back to a blank string if not found."""
        file_path = self.input_folder / filename

        if file_path.suffix.lower() == ".csv":
            with open(file_path, "r", encoding="utf-8") as f:
                title = f.readline()
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            title = ws.cell(row=1, column=1).value or ""
            wb.close()

        if "Month of" in title:
            return title.split("Month of", 1)[1].strip()
        return ""

       

    def load_employees(self, filename):
        df = self.load_excel(filename)
        df = df.fillna(0)

        employees = []
        for _, row in df.iterrows():
            data = {}
            for field, column in COLUMN_MAP.items():
                if column not in df.columns:
                    raise KeyError(
                        f"Expected column '{column}' not found in {filename}. "
                        f"Check the sheet's header row."
                    )
                data[field] = row[column]

            # epf/name/department must stay as text, not force-cast to float
            data["epf"] = str(data["epf"]).strip()
            data["name"] = str(data["name"]).strip()
            data["department"] = str(data["department"]).strip()

            employees.append(Employee(**data))

        return employees