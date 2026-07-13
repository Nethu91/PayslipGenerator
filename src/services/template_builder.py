from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class TemplateBuilder:

    def create(self):

        wb = Workbook()
        ws = wb.active
        ws.title = "Payslip"

        widths = {
            "A": 28,
            "B": 15,
            "C": 18,
            "D": 22,
            "E": 18,
            "F": 18,
        }

        for c, w in widths.items():
            ws.column_dimensions[c].width = w

        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        head_border = Border(
            left=medium,
            right=medium,
            top=medium,
            bottom=medium
        )

        blue = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        gray = PatternFill(
            fill_type="solid",
            fgColor="E7E6E6"
        )

        bold = Font(
            bold=True,
            size=11
        )

        title = Font(
            bold=True,
            size=16
        )

        center = Alignment(
            horizontal="center",
            vertical="center"
        )

        # ------------------------------------------------
        # TITLE
        # ------------------------------------------------

        ws.merge_cells("A1:F1")
        ws["A1"] = "SWISSTEK ALUMINIUM LTD"
        ws["A1"].font = title
        ws["A1"].alignment = center

        ws.merge_cells("A2:F2")
        ws["A2"] = "PAY SLIP"
        ws["A2"].font = Font(size=13, bold=True)
        ws["A2"].alignment = center

        # ------------------------------------------------
        # HEADER
        # ------------------------------------------------

        ws["A4"] = "EPF No"
        ws["D4"] = "Month"

        ws["A5"] = "Employee"
        ws["D5"] = "Department"

        ws["B4"] = ""
        ws["E4"] = ""

        ws["B5"] = ""
        ws["E5"] = ""

        # ------------------------------------------------
        # EARNINGS
        # ------------------------------------------------

        ws["A7"] = "EARNINGS"
        ws["E7"] = "AMOUNT"

        ws["A7"].fill = blue
        ws["E7"].fill = blue

        ws["A7"].font = bold
        ws["E7"].font = bold

        earnings = [

            "Basic Salary",

            "Holiday Pay",

            "Night Shift",

            "Saturday",

            "Sunday",

            "Poya",

            "Mercantile",

            "Over Time",

            "Double OT",

            "Chemical Allowance",

            "Export Allowance",

            "Attendance Incentive",

            "Production Incentive",

            "Team Leader Allowance",

            "Powder Incentive",

            "Extrusion Allowance",

            "Melt Allowance",

            "Tailoring Fee",

            "Refund ICEU",

            "Salary Arrears",

            "Gross Salary"

        ]

        row = 8

        for item in earnings:

            ws[f"A{row}"] = item

            ws[f"A{row}"].border = border
            ws[f"E{row}"].border = border

            row += 1

        # ------------------------------------------------
        # DEDUCTIONS
        # ------------------------------------------------

        start = 31

        ws[f"A{start}"] = "DEDUCTIONS"
        ws[f"E{start}"] = "AMOUNT"

        ws[f"A{start}"].fill = gray
        ws[f"E{start}"].fill = gray

        ws[f"A{start}"].font = bold
        ws[f"E{start}"].font = bold

        deductions = [

            "EPF 8%",

            "PAYE",

            "Meals",

            "Salary Advance",

            "Festival Advance",

            "Vision Care",

            "Other Deduction",

            "Welfare",

            "ICEU Fee",

            "Stamp Duty",

            "Total Deduction"

        ]

        row = start + 1

        for item in deductions:

            ws[f"A{row}"] = item

            ws[f"A{row}"].border = border
            ws[f"E{row}"].border = border

            row += 1

        # ------------------------------------------------

        ws["A44"] = "NET SALARY"
        ws["E44"] = ""

        ws["A44"].fill = blue
        ws["E44"].fill = blue

        ws["A44"].font = Font(size=12, bold=True)
        ws["E44"].font = Font(size=12, bold=True)

        # ------------------------------------------------

        ws["A47"] = "Employer Contribution"

        ws["A48"] = "EPF 12%"
        ws["A49"] = "ETF 3%"

        # ------------------------------------------------

        for r in range(4, 50):

            for c in range(1, 6):

                ws.cell(r, c).border = border

        wb.save("templates/payslip_template.xlsx")