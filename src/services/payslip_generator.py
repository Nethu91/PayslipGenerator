from pathlib import Path
from openpyxl import load_workbook


class PayslipGenerator:

    def __init__(self, template_path=None, output_folder=None):
        from src.config import TEMPLATE_FILE, EXCEL_DIR
        self.template = Path(template_path) if template_path else TEMPLATE_FILE
        self.output = Path(output_folder) if output_folder else EXCEL_DIR
        self.output.mkdir(parents=True, exist_ok=True)

    def _fill_sheet(self, ws, emp, pay_period):
        # ---------------- HEADER ----------------
        ws["B2"] = emp.epf
        ws["B3"] = emp.department
        ws["B5"] = emp.name
        ws["B6"] = pay_period

        # ---------------- EARNINGS ----------------
        ws["B10"] = emp.basic_salary
        ws["B11"] = emp.holiday_pay
        ws["B12"] = emp.no_pay
        ws["B13"] = emp.nopay_revise
        ws["B14"] = emp.salary_adjustment

        ws["B17"] = emp.night_shift
        ws["B18"] = emp.saturday_night
        ws["B19"] = emp.sunday_payment
        ws["B20"] = emp.poya_mercantile
        ws["B21"] = emp.mercantile_special_payment
        ws["B22"] = emp.refund_iceu_amount
        ws["B23"] = emp.ot_arrears
        ws["B24"] = emp.salary_arrears_deduction
        ws["B25"] = emp.over_time
        ws["B26"] = emp.double_time
        ws["B27"] = emp.powder_incentive
        ws["B28"] = emp.tailoring_fee
        ws["B29"] = emp.att_incentive
        ws["B30"] = emp.melt_incentive
        ws["B31"] = emp.team_leader_incentive
        ws["B32"] = emp.chemical_incentive
        ws["B33"] = emp.export_incentive
        ws["B34"] = emp.production_incentive

        # ---------------- DEDUCTIONS ----------------
        ws["B39"] = emp.salary_advance
        ws["B40"] = emp.festival_advance
        ws["B41"] = emp.vision_care
        ws["B42"] = emp.meals
        ws["B43"] = emp.welfare_society
        ws["B44"] = emp.other_deduction_tp
        ws["B45"] = emp.festival_advance_guarantors
        ws["B46"] = emp.iceu_member_fee
        ws["B47"] = emp.paye_tax
        ws["B48"] = emp.stamp_duty

        # ---------------- INFO ----------------
        ws["B55"] = emp.casual_leave
        ws["B56"] = emp.rate_per_not
        ws["B57"] = emp.rate_per_dot

    def generate_combined(self, employees, pay_period="", filename="All_Payslips.xlsx"):
        """One workbook, one sheet per employee, formulas preserved per sheet."""
        wb = load_workbook(self.template)
        template_ws = wb.active

        used_names = set()
        for i, emp in enumerate(employees):
            ws = template_ws if i == 0 else wb.copy_worksheet(template_ws)

            sheet_name = f"{emp.epf}_{emp.name}"[:31]  # Excel sheet-name limit
            base, n = sheet_name, 1
            while sheet_name in used_names:
                suffix = f"_{n}"
                sheet_name = base[: 31 - len(suffix)] + suffix
                n += 1
            used_names.add(sheet_name)
            ws.title = sheet_name

            self._fill_sheet(ws, emp, pay_period)

        output_path = self.output / filename

        # Remove read-only flag if the file already exists from a previous run
        if output_path.exists():
            import os
            import stat
            try:
                os.chmod(output_path, stat.S_IWRITE)
            except OSError:
                pass

        # Retry save in case of a transient lock (antivirus, OneDrive, etc.)
        import time
        last_err = None
        for attempt in range(3):
            try:
                wb.save(output_path)
                last_err = None
                break
            except PermissionError as e:
                last_err = e
                time.sleep(1)

        wb.close()

        if last_err is not None:
            raise PermissionError(
                f"Could not write to {output_path}.\n"
                f"Please check that the file is not open in Excel, "
                f"is not marked Read-only, and is not locked by antivirus/OneDrive.\n"
                f"Original error: {last_err}"
            )

        return output_path